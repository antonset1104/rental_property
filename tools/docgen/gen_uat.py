# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

NAVY = "1F3964"
GREEN = "006A4E"
GREY = "EFEFEF"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(bold=True, color=WHITE, size=10)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

# (ID, Module, Scenario, Steps, Expected, Priority)
CASES = [
    # Financial Report
    ("FR-01", "Financial Report", "Cetak Owners Statement dasar", "Pastikan ada invoice/bill posted ber-link properti. Jalankan wizard Owners Statement untuk periode berjalan.", "PDF tampil dengan 9 section; angka Income/Expense terisi dari GL.", "High"),
    ("FR-02", "Financial Report", "Actual vs Budget", "Buat Property Budget per akun untuk bulan uji. Cetak ulang Owners Statement.", "Kolom Budget & Variance & %Var pada I&E dan Performance Summary terisi.", "High"),
    ("FR-03", "Financial Report", "Multi-owner & remittance", "Isi 2 owner (60/40). Owner Remittance → Compute from Owners → Post.", "Alokasi 60/40; jurnal Dr Remittance/Cr Trust terbuat; Less Remittances muncul.", "High"),
    ("FR-04", "Financial Report", "Trust roll-forward", "Set Trust Bank Account; ada penerimaan & pembayaran ter-rekonsiliasi.", "Opening + Net Cash − Remittances = Closing Trust Balance konsisten.", "Medium"),
    ("FR-05", "Financial Report", "Security Deposit liability", "Buat Security Deposit → Mark as Held; tambah baris refund; Post Deductions/Refunds.", "Balance berkurang; kolom Sec Dep Bal di Tenant Balances terisi.", "Medium"),
    ("FR-06", "Financial Report", "Integrasi Analytic", "Set analytic account properti. Posting invoice ber-link properti. Buka 'Analytic Items'.", "Baris income/expense punya analytic_distribution; muncul di Analytic Items.", "Medium"),
    ("FR-07", "Financial Report", "Aged Arrears", "Biarkan beberapa invoice jatuh tempo umur berbeda. Cetak laporan.", "Tunggakan terbagi ke bucket Current/1/2/3/4+ bulan dengan total benar.", "Medium"),
    # GTO & Meter
    ("GT-01", "GTO & Meter", "GTO Higher-of", "Kontrak GTO higher_of, %=5, base=10jt, turnover=300jt. Compute.", "percentage_rent=15jt; billable=5jt; invoice 5jt.", "High"),
    ("GT-02", "GTO & Meter", "GTO Base+Overage", "base_plus, %=5, breakpoint=200jt, turnover=300jt.", "billable=(300−200)jt×5%=5jt.", "Medium"),
    ("GT-03", "GTO & Meter", "GTO Pure %", "pure, %=8, turnover=100jt.", "billable=8jt.", "Low"),
    ("MT-01", "GTO & Meter", "Meter reading & recharge", "Tarif=1.500/kWh, previous=100, current=350. Create Recharge Invoice.", "consumption=250; amount=375.000; invoice qty 250×1.500.", "High"),
    ("MT-02", "GTO & Meter", "Previous reading otomatis", "Buat reading kedua pada meter sama.", "previous_reading terisi otomatis dari current terakhir.", "Medium"),
    ("GT-04", "GTO & Meter", "Aliran ke laporan", "Posting invoice GTO/meter, cetak Owners Statement.", "Pendapatan muncul di Income Owners Statement.", "Medium"),
    # Guarantee
    ("GU-01", "Guarantee", "Buat & aktifkan jaminan", "Buat guarantee expiry 60 hari, Activate.", "State=Active; days_to_expiry≈60.", "High"),
    ("GU-02", "Guarantee", "Expiring soon", "Set expiry 10 hari (reminder 30). Jalankan cron.", "is_expiring=true; activity reminder terjadwal.", "Medium"),
    ("GU-03", "Guarantee", "Auto-expire", "Set expiry kemarin. Jalankan cron.", "State=Expired; ada log pesan.", "Medium"),
    ("GU-04", "Guarantee", "Release/Claim", "Klik Release lalu Claim pada jaminan berbeda.", "State berubah sesuai aksi.", "Low"),
    # Casual Leasing
    ("CL-01", "Casual Leasing", "Tarif per hari", "Periode 5 hari, Per Day, tarif=200rb.", "duration_days=5; total=1jt.", "High"),
    ("CL-02", "Casual Leasing", "Tarif per minggu", "Periode 10 hari, Per Week, tarif=1jt.", "quantity=ceil(10/7)=2; total=2jt.", "Medium"),
    ("CL-03", "Casual Leasing", "Buat invoice", "Confirm lalu Create Invoice.", "out_invoice sejumlah total; state=Active; ber-link properti.", "High"),
    # Handover
    ("HO-01", "Handover", "Move-in checklist", "Buat handover Move-in, Load Default Checklist.", "5 item checklist Move-in muncul.", "High"),
    ("HO-02", "Handover", "Progress", "Centang 3 dari 6 item.", "progress=50%.", "Low"),
    ("HO-03", "Handover", "Lifecycle", "Start lalu Complete.", "State=Completed; actual_date terisi.", "Medium"),
    ("HO-04", "Handover", "Fit-out fields", "Tipe Fit-out: isi kontraktor & bond.", "Field fit-out tampil hanya untuk Fit-out.", "Low"),
    # Portal
    ("PT-01", "Tenant Portal", "Akses portal tenant", "Login sebagai tenant portal yang punya kontrak.", "Kartu Contracts tampil dengan jumlah benar.", "High"),
    ("PT-02", "Tenant Portal", "Isolasi data", "Tenant A akses kontrak tenant B (ubah id URL).", "Diarahkan ke /my (akses ditolak).", "High"),
    ("PT-03", "Tenant Portal", "Detail & invoice", "Buka detail kontrak, klik View My Invoices.", "Detail tampil; daftar invoice tenant muncul.", "Medium"),
    # CORETAX
    ("CT-01", "CORETAX", "Ekspor Faktur PK", "Set TIN company & buyer. Posting invoice. Export Faktur Keluaran.", "XML TaxInvoiceBulk terunduh; struktur v1.4; invoice ditandai exported.", "High"),
    ("CT-02", "CORETAX", "Retur PM", "Vendor credit note + Original Faktur No. Export Retur PM.", "InputTaxInvoiceReturn berisi Rows + FooterRow total.", "Medium"),
    ("CT-03", "CORETAX", "L9 Penyusutan", "Isi register L9, Tax Year. Export.", "DepreciationAmortization; nilai desimal utuh.", "Medium"),
    ("CT-04", "CORETAX", "Validasi tanpa TIN", "Hapus TIN company, coba Export PK.", "Muncul peringatan agar set TIN/NPWP.", "Low"),
    # Purchase
    ("PU-01", "Purchase", "PO ber-properti → bill", "Buat PO Property terisi, konfirmasi, Create Bill, post.", "Vendor bill ber-link properti; muncul di Payment Details & analytic.", "High"),
    ("PU-02", "Purchase", "PO dari maintenance", "Maintenance request (vendor & produk) → Create Purchase Order.", "PO terbuat dengan baris produk request; maintenance_request_id terisi.", "Medium"),
    # CRM
    ("CR-01", "CRM", "Create Lease Contract", "Lead (partner & property) → Create Lease Contract.", "Form tenancy terbuka dengan property & tenant ter-default.", "High"),
    ("CR-02", "CRM", "Smart button kontrak", "Lead dengan partner yang punya kontrak.", "Smart button menampilkan jumlah & daftar kontrak.", "Low"),
    # Project
    ("PJ-01", "Project (Fit-out)", "Buat project fit-out", "Handover Fit-out 6 item → Create Fit-out Project.", "project + 6 task terbuat; project_id terisi.", "High"),
    ("PJ-02", "Project (Fit-out)", "Lihat tasks", "Klik Fit-out Tasks.", "Daftar 6 task pada project tampil.", "Low"),
    # Documents
    ("DM-01", "Documents", "Buat folder & sync", "Properti dengan attachment → Open / Sync Documents.", "Folder terbuat; dokumen muncul; documents_count > 0.", "Medium"),
    # Fixed Assets
    ("AS-01", "Fixed Assets", "Board garis lurus", "Aset 10jt, salvage 0, linear, jumlah=5, periode=12. Compute Depreciation.", "5 baris @ 2jt; akumulasi akhir=10jt; remaining akhir=0.", "High"),
    ("AS-02", "Fixed Assets", "Posting penyusutan", "Confirm; tanggal baris ≤ hari ini; Post Due Depreciation.", "Jurnal Dr Beban/Cr Akumulasi posted; ber-tag properti (analytic).", "High"),
    ("AS-03", "Fixed Assets", "Saldo menurun", "Aset declining, factor=2, jumlah=5. Compute.", "Nominal menurun tiap periode; baris terakhir menutup ke salvage.", "Medium"),
    ("AS-04", "Fixed Assets", "Revaluation naik", "Tambah revaluation +5jt → Post Revaluations.", "Jurnal Dr Aset/Cr Reserve 5jt; gross naik; board recompute.", "High"),
    ("AS-05", "Fixed Assets", "Revaluation turun", "Tambah revaluation −2jt → Post Revaluations.", "Jurnal Dr Reserve/Cr Aset 2jt; gross turun; board recompute.", "Medium"),
    ("AS-06", "Fixed Assets", "Sync CORETAX L9", "Klik Sync CORETAX L9 (coretax terpasang).", "Entri coretax.asset.depreciation terbuat dgn nilai perolehan & penyusutan tahun berjalan.", "Low"),
    # Documents (Community)
    ("DC-01", "Documents CE", "Unggah lampiran", "Lampirkan 2 dokumen pada properti via tab Documents.", "documents_count=2; tersimpan sebagai ir.attachment milik properti.", "Low"),
    # Owner Portal
    ("OP-01", "Owner Portal", "Akses pemilik", "Login sebagai owner portal yang memiliki properti.", "Daftar Properties hanya menampilkan properti miliknya.", "High"),
    ("OP-02", "Owner Portal", "Isolasi data", "Owner A coba akses properti owner B via id URL.", "Akses ditolak / diarahkan ke /my.", "High"),
    ("OP-03", "Owner Portal", "Lihat remittance", "Buka detail properti dengan remittance posted.", "Daftar remittance tampil dengan jumlah benar.", "Medium"),
    # CAM
    ("CM-01", "CAM / Service Charge", "Apportion per area", "Pool=120jt; 3 tenant area 100/200/300 m². Apportion.", "Alokasi 20/40/60 jt sesuai proporsi area.", "High"),
    ("CM-02", "CAM / Service Charge", "Invoice service charge", "Klik Create Invoices.", "out_invoice per tenant ber-link properti; muncul di Owners Statement & analytic.", "High"),
    ("CM-03", "CAM / Service Charge", "Budget vs actual", "Isi actual berbeda dari budget.", "Variance terhitung untuk rekonsiliasi.", "Medium"),
    # Rent Escalation
    ("RE-01", "Rent Escalation", "Eskalasi %", "Sewa 10jt, 10%/tahun, due hari ini. Jalankan cron.", "Sewa naik ke 11jt; log mencatat 10jt→11jt.", "High"),
    ("RE-02", "Rent Escalation", "Eskalasi amount", "Escalation +500rb. Jalankan cron.", "Sewa naik 500rb; log tercatat.", "Medium"),
    ("RE-03", "Rent Escalation", "Belum jatuh tempo", "Tanggal due di masa depan. Jalankan cron.", "Tidak ada perubahan sewa.", "Low"),
    # Dashboard
    ("DB-01", "KPI Dashboard", "Hitung KPI", "Dengan kontrak & invoice ada, buka Dashboard.", "NOI=income−expense; arrears=invoice overdue; collection rate konsisten.", "High"),
    ("DB-02", "KPI Dashboard", "Lease expiring", "Ada kontrak berakhir dalam 12 bulan.", "Counter leases-expiring menampilkan jumlah benar.", "Medium"),
    # Dunning
    ("DN-01", "Dunning", "Naik level", "Invoice overdue melewati ambang level 1. Jalankan cron.", "Dunning level=1; email reminder terkirim.", "High"),
    ("DN-02", "Dunning", "Late fee", "Aktifkan late fee pada level. Jalankan cron.", "Denda ditambahkan sesuai konfigurasi.", "Medium"),
    ("DN-03", "Dunning", "Lunas berhenti", "Bayar invoice. Jalankan cron.", "Tidak ada reminder lanjutan.", "Medium"),
    # Insurance
    ("IN-01", "Insurance", "Buat polis", "Buat polis expiry 30 hari ke depan.", "Polis tersimpan; days-to-expiry terhitung.", "Medium"),
    ("IN-02", "Insurance", "Reminder", "Expiry dalam lead window. Jalankan cron.", "Activity reminder terjadwal pada penanggung jawab.", "Medium"),
    # Lease Expiry
    ("LE-01", "Lease Expiry", "Filter 30 hari", "Kontrak berakhir dalam 20 hari.", "Muncul pada filter 30 hari.", "High"),
    ("LE-02", "Lease Expiry", "Reminder renewal", "Jalankan cron.", "Activity reminder renewal terjadwal.", "Medium"),
    # Vacancy
    ("VC-01", "Vacancy", "Kelompok status", "Beberapa properti status berbeda.", "Board mengelompokkan benar sesuai stage.", "Medium"),
    ("VC-02", "Vacancy", "Filter available", "Pilih filter Available.", "Hanya properti available tampil.", "Low"),
    # PPM
    ("PP-01", "PPM", "Generate request", "PPM plan due hari ini. Jalankan cron.", "maintenance.request baru terbuat sesuai plan.", "High"),
    ("PP-02", "PPM", "Interval berikut", "Setelah generate, cek next date.", "Tanggal jadwal berikut bergeser sesuai interval.", "Medium"),
    # Parking
    ("PK-01", "Parking", "Alokasi bay", "Tetapkan bay ke tenant.", "Bay berstatus allocated; terkait tenant/kontrak.", "Medium"),
    ("PK-02", "Parking", "Invoice parkir", "Buat invoice parkir.", "out_invoice ber-link properti; muncul di Owners Statement.", "Medium"),
    # Valuation
    ("VL-01", "Valuation", "Catat valuasi", "Buat 2 valuasi tanggal berbeda.", "Latest value = valuasi tanggal terbaru.", "Low"),
    # Access
    ("AC-01", "Access", "Terbitkan kartu", "Buat kartu akses untuk tenant.", "Kartu tercatat & terkait tenant.", "Low"),
    ("AC-02", "Access", "Log tamu", "Catat check-in lalu check-out tamu.", "Waktu check-in/out tersimpan.", "Low"),
    # E-sign
    ("ES-01", "E-Signature", "Kirim permintaan", "Klik Request Signature pada kontrak.", "Status=Sent; sequence terbuat; email/log tercatat.", "Medium"),
    ("ES-02", "E-Signature", "Tandai signed", "Klik Mark Signed.", "Status=Signed; tanggal tanda tangan tersimpan.", "Low"),
    # End-to-end
    ("E2E-1", "End-to-End", "Onboarding tenant", "Lead → Create Lease Contract → aktifkan → Handover Move-in → Guarantee.", "Kontrak aktif; handover selesai; jaminan Active; tenant bisa login portal.", "High"),
    ("E2E-2", "End-to-End", "Billing bulanan", "Invoice sewa + Meter recharge + GTO turnover diposting.", "Semua pendapatan ber-link properti & analytic; muncul di Owners Statement.", "High"),
    ("E2E-3", "End-to-End", "Pengeluaran", "Maintenance → Create PO → terima → vendor bill posted.", "Beban masuk Payment Details & analytic properti.", "High"),
    ("E2E-4", "End-to-End", "Tutup periode", "Cetak Owners Statement; Owner Remittance; ekspor CORETAX Faktur PK.", "Laporan konsisten; remittance terposting; XML e-Faktur terunduh.", "High"),
    ("E2E-5", "End-to-End", "Move-out", "Handover Move-out + make-good; Security Deposit refund; kontrak ditutup.", "Deposit terselesaikan; Sec Dep Bal nol; kontrak closed.", "Medium"),
]

wb = openpyxl.Workbook()

# ---------------- Test Cases sheet ----------------
ws = wb.active
ws.title = "Test Cases"
headers = ["ID", "Modul", "Skenario", "Langkah Pengujian", "Hasil yang Diharapkan",
           "Prioritas", "Status", "Penguji", "Tanggal Uji", "Hasil Aktual / Catatan"]
widths = [9, 18, 26, 42, 42, 10, 12, 16, 13, 40]
ws.append(headers)
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = hdr_font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = border
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:J1"

for row in CASES:
    rid, mod, sc, st, exp, prio = row
    ws.append([rid, mod, sc, st, exp, prio, "Not Run", "", "", ""])

# style data rows
for r in range(2, len(CASES) + 2):
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        cell.alignment = wrap if c in (3, 4, 5, 10) else Alignment(vertical="top", horizontal="center" if c in (1, 6, 7, 9) else "left")
    if r % 2 == 0:
        for c in range(1, len(headers) + 1):
            if ws.cell(row=r, column=c).fill.fgColor.rgb in (None, "00000000"):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="F5F8FC")

# Data validation: Status dropdown
status_dv = DataValidation(type="list", formula1='"Not Run,Pass,Fail,Blocked,N/A"', allow_blank=True)
ws.add_data_validation(status_dv)
status_dv.add(f"G2:G{len(CASES) + 1}")
prio_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws.add_data_validation(prio_dv)
prio_dv.add(f"F2:F{len(CASES) + 1}")

# Conditional formatting for Status
from openpyxl.formatting.rule import CellIsRule
green_fill = PatternFill("solid", fgColor="C6EFCE")
red_fill = PatternFill("solid", fgColor="FFC7CE")
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
grey_fill = PatternFill("solid", fgColor="E0E0E0")
rng = f"G2:G{len(CASES) + 1}"
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Pass"'], fill=green_fill))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Fail"'], fill=red_fill))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Blocked"'], fill=yellow_fill))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Not Run"'], fill=grey_fill))

n = len(CASES)
last = n + 1

# ---------------- Summary sheet ----------------
sm = wb.create_sheet("Summary")
sm.column_dimensions["A"].width = 22
sm.column_dimensions["B"].width = 14
sm["A1"] = "UAT SUMMARY"
sm["A1"].font = Font(bold=True, size=14, color=NAVY)
rows_sum = [
    ("Total Test Cases", f'=COUNTA(\'Test Cases\'!A2:A{last})'),
    ("Pass", f'=COUNTIF(\'Test Cases\'!G2:G{last},"Pass")'),
    ("Fail", f'=COUNTIF(\'Test Cases\'!G2:G{last},"Fail")'),
    ("Blocked", f'=COUNTIF(\'Test Cases\'!G2:G{last},"Blocked")'),
    ("N/A", f'=COUNTIF(\'Test Cases\'!G2:G{last},"N/A")'),
    ("Not Run", f'=COUNTIF(\'Test Cases\'!G2:G{last},"Not Run")'),
    ("% Pass", f'=IF(B3+B4=0,0,B3/(B3+B4))'),
]
r0 = 3
for i, (label, formula) in enumerate(rows_sum):
    rr = r0 + i
    sm.cell(row=rr, column=1, value=label).font = Font(bold=True)
    ccell = sm.cell(row=rr, column=2, value=formula)
    ccell.alignment = center
    sm.cell(row=rr, column=1).border = border
    ccell.border = border
sm.cell(row=r0 + 6, column=2).number_format = "0.0%"

# per-module breakdown
sm.cell(row=r0 + 8, column=1, value="Per Modul").font = Font(bold=True, size=12, color=GREEN)
hdr2 = ["Modul", "Total", "Pass", "Fail", "Not Run"]
for j, htxt in enumerate(hdr2, 1):
    c = sm.cell(row=r0 + 9, column=j, value=htxt)
    c.fill = PatternFill("solid", fgColor=GREEN); c.font = hdr_font; c.border = border
    c.alignment = center
mods = []
for row in CASES:
    if row[1] not in mods:
        mods.append(row[1])
for k, mod in enumerate(mods):
    rr = r0 + 10 + k
    sm.cell(row=rr, column=1, value=mod).border = border
    sm.cell(row=rr, column=2, value=f'=COUNTIF(\'Test Cases\'!B2:B{last},A{rr})').border = border
    sm.cell(row=rr, column=3, value=f'=COUNTIFS(\'Test Cases\'!B2:B{last},A{rr},\'Test Cases\'!G2:G{last},"Pass")').border = border
    sm.cell(row=rr, column=4, value=f'=COUNTIFS(\'Test Cases\'!B2:B{last},A{rr},\'Test Cases\'!G2:G{last},"Fail")').border = border
    sm.cell(row=rr, column=5, value=f'=COUNTIFS(\'Test Cases\'!B2:B{last},A{rr},\'Test Cases\'!G2:G{last},"Not Run")').border = border
    for col in range(2, 6):
        sm.cell(row=rr, column=col).alignment = center
    for col in range(2, 6):
        sm.column_dimensions[get_column_letter(col)].width = 10

# ---------------- Sign-off sheet ----------------
so = wb.create_sheet("Sign-off")
so.column_dimensions["A"].width = 24
so.column_dimensions["B"].width = 40
so["A1"] = "UAT SIGN-OFF"
so["A1"].font = Font(bold=True, size=14, color=NAVY)
info = [
    ("Project", "Kustomisasi Property Management Odoo 19 (rental_management companions)"),
    ("Versi Modul", "19.0.1.0.0"),
    ("Environment", "Staging Odoo 19"),
    ("Tanggal Mulai UAT", ""),
    ("Tanggal Selesai UAT", ""),
    ("Diuji oleh (Key User)", ""),
    ("Disetujui oleh (PM/SME)", ""),
    ("Catatan", ""),
]
for i, (k, v) in enumerate(info):
    rr = 3 + i
    so.cell(row=rr, column=1, value=k).font = Font(bold=True)
    so.cell(row=rr, column=1).border = border
    so.cell(row=rr, column=2, value=v).border = border

wb.save("../../docs/UAT_Tracker_Custom_Modules.xlsx")
print("SAVED OK; total cases:", n)
