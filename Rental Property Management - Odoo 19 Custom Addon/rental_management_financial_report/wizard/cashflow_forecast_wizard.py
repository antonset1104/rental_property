# -*- coding: utf-8 -*-
import base64
import io
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from odoo import api, fields, models, _
from odoo.exceptions import UserError


class PropertyCashflowForecastWizard(models.TransientModel):
    _name = 'property.cashflow.forecast.wizard'
    _description = 'Proyeksi Arus Kas Bergulir 12 Bulan (12-Month Rolling Cash Flow)'

    start_date = fields.Date(string='Bulan Awal Proyeksi', default=lambda s: fields.Date.today().replace(day=1),
                             required=True)
    company_id = fields.Many2one('res.company', string='Perusahaan / PT',
                                 default=lambda s: s.env.company, required=True)
    property_id = fields.Many2one('property.details', string='Filter Gedung (Opsional)',
                                  help="Kosongkan untuk menghitung seluruh portofolio gedung.")

    excel_file = fields.Binary(string='File Excel', readonly=True)
    excel_filename = fields.Char(string='Nama File Excel', readonly=True)

    def _get_forecast_data(self):
        self.ensure_one()
        s_date = self.start_date.replace(day=1)
        months = []
        for i in range(12):
            m_date = s_date + relativedelta(months=i)
            m_end = m_date + relativedelta(months=1, days=-1)
            months.append({
                'month_idx': i + 1,
                'label': m_date.strftime('%b %Y'),
                'date_from': m_date,
                'date_to': m_end,
                'inflow_rent': 0.0,
                'inflow_service': 0.0,
                'inflow_utility': 0.0,
                'inflow_deposit': 0.0,
                'total_inflow': 0.0,
                'outflow_vendor': 0.0,
                'outflow_maintenance': 0.0,
                'outflow_tax_pbb': 0.0,
                'outflow_deposit_refund': 0.0,
                'total_outflow': 0.0,
                'net_cashflow': 0.0,
                'cumulative_cashflow': 0.0,
            })

        # Calculate Inflows from active tenancies
        t_domain = [
            ('company_id', '=', self.company_id.id),
            ('state', 'in', ('draft', 'in_progress', 'close')),
        ]
        if self.property_id:
            t_domain.append(('property_id', '=', self.property_id.id))

        tenancies = self.env['tenancy.details'].search(t_domain)
        for t in tenancies:
            t_rent = t.total_rent or 0.0
            # Distribute over months
            for m in months:
                if t.start_date and t.end_date:
                    if t.start_date <= m['date_to'] and t.end_date >= m['date_from']:
                        # monthly rent estimate
                        m_rent = t_rent / 12.0 if t_rent > 0 else 5000000.0
                        m['inflow_rent'] += m_rent
                        m['inflow_service'] += m_rent * 0.15
                        m['inflow_utility'] += m_rent * 0.10

        # Cumulative calculation
        cum = 0.0
        for m in months:
            # Outflow estimates based on property operations
            m['outflow_vendor'] = m['inflow_service'] * 0.40
            m['outflow_maintenance'] = m['inflow_rent'] * 0.08
            m['outflow_tax_pbb'] = m['inflow_rent'] * 0.05
            
            m['total_inflow'] = m['inflow_rent'] + m['inflow_service'] + m['inflow_utility'] + m['inflow_deposit']
            m['total_outflow'] = m['outflow_vendor'] + m['outflow_maintenance'] + m['outflow_tax_pbb'] + m['outflow_deposit_refund']
            m['net_cashflow'] = m['total_inflow'] - m['total_outflow']
            cum += m['net_cashflow']
            m['cumulative_cashflow'] = cum

        return {
            'company_name': self.company_id.name,
            'property_name': self.property_id.name if self.property_id else 'Konsolidasian Seluruh Gedung',
            'start_month': s_date.strftime('%B %Y'),
            'months': months,
            'grand_total_inflow': sum(m['total_inflow'] for m in months),
            'grand_total_outflow': sum(m['total_outflow'] for m in months),
            'grand_net_cashflow': sum(m['net_cashflow'] for m in months),
        }

    def action_print_pdf(self):
        self.ensure_one()
        return self.env.ref('rental_management_financial_report.action_report_cashflow_forecast').report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        data = self._get_forecast_data()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proyeksi Arus Kas 12 Bulan"

        header_fill = PatternFill(start_color="1F3964", end_color="1F3964", fill_type="solid")
        section_fill = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="1F3964")
        bold_font = Font(name="Calibri", size=10, bold=True)
        normal_font = Font(name="Calibri", size=10)
        border_thin = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        ws.append(["PROYEKSI ARUS KAS BERGULIR 12 BULAN (ROLLING CASH FLOW FORECAST)"])
        ws.cell(row=1, column=1).font = title_font
        ws.append([f"Entitas: {data['company_name']} | Portofolio: {data['property_name']} | Periode: {data['start_month']}"])
        ws.append([])

        headers = ["Komponen Arus Kas (IDR)"] + [m['label'] for m in data['months']] + ["Total 12 Bulan"]
        ws.append(headers)
        hdr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=hdr_row, column=col_idx)
            c.fill = header_fill
            c.font = header_font
            c.border = border_thin

        # 1. Inflows
        rows = [
            ("A. ARUS KAS MASUK (INFLOWS)", ["" for _ in data['months']], ""),
            ("  1. Pendapatan Sewa Pokok", [m['inflow_rent'] for m in data['months']], sum(m['inflow_rent'] for m in data['months'])),
            ("  2. Service Charge Terjadwal", [m['inflow_service'] for m in data['months']], sum(m['inflow_service'] for m in data['months'])),
            ("  3. Tagihan Utilitas (Listrik & Air)", [m['inflow_utility'] for m in data['months']], sum(m['inflow_utility'] for m in data['months'])),
            ("  TOTAL ARUS KAS MASUK", [m['total_inflow'] for m in data['months']], data['grand_total_inflow']),
            ("B. ARUS KAS KELUAR (OUTFLOWS)", ["" for _ in data['months']], ""),
            ("  1. Kontrak Vendor (Cleaning & Security)", [m['outflow_vendor'] for m in data['months']], sum(m['outflow_vendor'] for m in data['months'])),
            ("  2. Pemeliharaan Rutin & MEP", [m['outflow_maintenance'] for m in data['months']], sum(m['outflow_maintenance'] for m in data['months'])),
            ("  3. Estimasi Pajak & PBB", [m['outflow_tax_pbb'] for m in data['months']], sum(m['outflow_tax_pbb'] for m in data['months'])),
            ("  TOTAL ARUS KAS KELUAR", [m['total_outflow'] for m in data['months']], data['grand_total_outflow']),
            ("C. NET CASH FLOW BULANAN", [m['net_cashflow'] for m in data['months']], data['grand_net_cashflow']),
            ("D. ARUS KAS KUMULATIF", [m['cumulative_cashflow'] for m in data['months']], data['grand_net_cashflow']),
        ]

        for item, m_vals, total_val in rows:
            is_section = item.startswith(('A.', 'B.', 'TOTAL', 'C.', 'D.'))
            row_data = [item] + list(m_vals) + [total_val]
            ws.append(row_data)
            curr_r = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=curr_r, column=col_idx)
                c.border = border_thin
                if is_section:
                    c.font = bold_font
                    if item.startswith(('A.', 'B.')):
                        c.fill = section_fill
                else:
                    c.font = normal_font

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = output.read()
        output.close()

        filename = f"Cashflow_Forecast_{fields.Date.today().strftime('%Y%m')}.xlsx"
        self.write({
            'excel_file': base64.b64encode(file_data),
            'excel_filename': filename,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=property.cashflow.forecast.wizard&id={self.id}&field=excel_file&download=true&filename={filename}',
            'target': 'self',
        }


class ReportCashflowForecast(models.AbstractModel):
    _name = 'report.rental_management_financial_report.report_cashflow_forecast'
    _description = 'Laporan Proyeksi Arus Kas PDF'

    @api.model
    def _get_report_values(self, docids, data=None):
        wizard = self.env['property.cashflow.forecast.wizard'].browse(docids)
        return {
            'doc_ids': docids,
            'doc_model': 'property.cashflow.forecast.wizard',
            'docs': wizard,
            'report_data': wizard._get_forecast_data(),
        }
