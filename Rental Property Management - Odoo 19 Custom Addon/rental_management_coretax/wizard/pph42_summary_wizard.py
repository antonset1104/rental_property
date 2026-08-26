# -*- coding: utf-8 -*-
import base64
import io
from datetime import date
from odoo import api, fields, models, _
from odoo.exceptions import UserError


class PropertyPPh42SummaryWizard(models.TransientModel):
    _name = 'property.pph42.summary.wizard'
    _description = 'Rekapitulasi Tahunan PPh Final Pasal 4(2) & e-Bupot'

    year = fields.Selection(
        [(str(y), str(y)) for y in range(2022, 2035)],
        string='Tahun Pajak (Tahun Fiskal)',
        default=lambda s: str(fields.Date.today().year),
        required=True
    )
    company_id = fields.Many2one('res.company', string='Perusahaan / PT',
                                 default=lambda s: s.env.company, required=True)
    partner_id = fields.Many2one('res.partner', string='Penyewa / Tenant (Opsional)',
                                 help="Biarkan kosong untuk merekapitulasi seluruh tenant.")
    status_filter = fields.Selection([
        ('all', 'Semua Status e-Bupot'),
        ('received', 'Sudah Diterima (Received)'),
        ('pending', 'Menunggu Bukti Potong (Pending / Tertunggak)'),
    ], string='Status Bukti Potong', default='all', required=True)

    excel_file = fields.Binary(string='File Excel', readonly=True)
    excel_filename = fields.Char(string='Nama File Excel', readonly=True)

    def _get_data(self):
        self.ensure_one()
        y = int(self.year)
        d_from = date(y, 1, 1)
        d_to = date(y, 12, 31)

        domain = [
            ('move_type', 'in', ('out_invoice', 'out_refund')),
            ('state', '=', 'posted'),
            ('invoice_date', '>=', d_from),
            ('invoice_date', '<=', d_to),
            ('company_id', '=', self.company_id.id),
            ('is_pph42_applicable', '=', True),
        ]
        if self.partner_id:
            domain.append(('partner_id', '=', self.partner_id.id))
        if self.status_filter == 'received':
            domain.append(('pph42_ebupot_status', '=', 'received'))
        elif self.status_filter == 'pending':
            domain.append(('pph42_ebupot_status', 'in', ('none', 'pending')))

        moves = self.env['account.move'].search(domain, order='invoice_date, id')

        partners_data = {}
        total_dpp_all = 0.0
        total_pph_all = 0.0
        total_received_pph = 0.0
        total_pending_pph = 0.0

        for m in moves:
            p = m.partner_id
            pid = p.id
            if pid not in partners_data:
                partners_data[pid] = {
                    'partner_name': p.name or 'Tanpa Nama',
                    'npwp': p.vat or '-',
                    'total_dpp': 0.0,
                    'total_pph': 0.0,
                    'received_pph': 0.0,
                    'pending_pph': 0.0,
                    'invoices': [],
                }
            dpp = m.amount_untaxed if m.move_type == 'out_invoice' else -m.amount_untaxed
            pph = m.pph42_amount if m.move_type == 'out_invoice' else -m.pph42_amount

            partners_data[pid]['total_dpp'] += dpp
            partners_data[pid]['total_pph'] += pph
            total_dpp_all += dpp
            total_pph_all += pph

            if m.pph42_ebupot_status == 'received':
                partners_data[pid]['received_pph'] += pph
                total_received_pph += pph
            else:
                partners_data[pid]['pending_pph'] += pph
                total_pending_pph += pph

            partners_data[pid]['invoices'].append({
                'invoice_number': m.name,
                'invoice_date': m.invoice_date,
                'dpp': dpp,
                'pph': pph,
                'bupot_no': m.pph42_ebupot_number or '-',
                'bupot_date': m.pph42_ebupot_date or False,
                'status': m.pph42_ebupot_status,
            })

        return {
            'year': self.year,
            'company_name': self.company_id.name,
            'company_vat': self.company_id.vat or '-',
            'partners': list(partners_data.values()),
            'total_dpp_all': total_dpp_all,
            'total_pph_all': total_pph_all,
            'total_received_pph': total_received_pph,
            'total_pending_pph': total_pending_pph,
            'move_count': len(moves),
        }

    def action_print_pdf(self):
        self.ensure_one()
        return self.env.ref('rental_management_coretax.action_report_pph42_annual_summary').report_action(self)

    def action_export_excel(self):
        self.ensure_one()
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        data = self._get_data()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"PPh 4(2) Rekap {self.year}"

        header_fill = PatternFill(start_color="1F3964", end_color="1F3964", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="1F3964")
        bold_font = Font(name="Calibri", size=10, bold=True)
        normal_font = Font(name="Calibri", size=10)
        border_thin = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        ws.append(["REKAPITULASI TAHUNAN BUKTI POTONG PPH FINAL PASAL 4 AYAT (2) SEWA GEDUNG"])
        ws.cell(row=1, column=1).font = title_font
        ws.append([f"Entitas: {data['company_name']} | NPWP: {data['company_vat']} | Tahun Pajak: {data['year']}"])
        ws.append([])

        ws.append(["Ringkasan Rekapitulasi", "Nominal (IDR)"])
        ws.append(["Total DPP Sewa Tanah & Bangunan", data['total_dpp_all']])
        ws.append(["Total Estimasi PPh Final 4(2) 10%", data['total_pph_all']])
        ws.append(["Realisasi e-Bupot Diterima", data['total_received_pph']])
        ws.append(["e-Bupot Belum Diserahkan (Pending)", data['total_pending_pph']])
        ws.append([])

        for r in range(4, 9):
            for c in range(1, 3):
                cell = ws.cell(row=r, column=c)
                cell.border = border_thin
                if r == 4:
                    cell.fill = header_fill
                    cell.font = header_font
                elif c == 1:
                    cell.font = bold_font

        headers = ["No.", "Nama Penyewa (Tenant)", "NPWP Tenant", "Total DPP Sewa (IDR)", "PPh 4(2) Dipotong 10% (IDR)", "e-Bupot Diterima (IDR)", "Pending e-Bupot (IDR)", "Status"]
        ws.append(headers)
        hdr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=hdr_row, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border_thin

        for idx, p in enumerate(data['partners'], start=1):
            status_txt = "Lengkap" if p['pending_pph'] == 0 else f"Tertunggak ({len([i for i in p['invoices'] if i['status'] != 'received'])} Faktur)"
            row = [
                idx,
                p['partner_name'],
                p['npwp'],
                p['total_dpp'],
                p['total_pph'],
                p['received_pph'],
                p['pending_pph'],
                status_txt
            ]
            ws.append(row)
            curr_r = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=curr_r, column=col_idx)
                cell.border = border_thin
                cell.font = normal_font

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = output.read()
        output.close()

        filename = f"Rekap_PPh42_{data['year']}_{self.company_id.name.replace(' ', '_')}.xlsx"
        self.write({
            'excel_file': base64.b64encode(file_data),
            'excel_filename': filename,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model=property.pph42.summary.wizard&id={self.id}&field=excel_file&download=true&filename={filename}',
            'target': 'self',
        }


class ReportPPh42AnnualSummary(models.AbstractModel):
    _name = 'report.rental_management_coretax.report_pph42_annual_summary'
    _description = 'Laporan Rekapitulasi Tahunan PPh 4(2) PDF'

    @api.model
    def _get_report_values(self, docids, data=None):
        wizard = self.env['property.pph42.summary.wizard'].browse(docids)
        return {
            'doc_ids': docids,
            'doc_model': 'property.pph42.summary.wizard',
            'docs': wizard,
            'report_data': wizard._get_data(),
        }
